"""
Paper list export — CSV / XLSX of the papers table, usable at the Filter
stage (before extraction has run) or the Sources stage (after).

Deliberately independent of the review-export path in core/exporters.py:
this exports the DATA (one row per paper, with whatever extraction fields
exist yet), not the generated review. Extraction columns are simply blank
for any paper that hasn't been extracted yet — this is meant to work at
either stage without the caller needing to know which one it's in.
"""
import csv
import io

HEADERS = [
    "#", "Title", "Authors", "Year", "Venue", "Source", "URL", "Included", "Abstract",
    "Method", "Finding", "Dataset", "Metrics", "Limitation", "Contribution", "Relevance", "Concepts",
]

# Column widths for the xlsx sheet, same order as HEADERS.
_WIDTHS = [4, 44, 24, 6, 18, 14, 32, 9, 46, 26, 30, 18, 22, 26, 30, 30, 24]


def _rows(papers: list[dict], extractions_by_idx: dict, included_map: dict) -> list[list]:
    rows = []
    for i, p in enumerate(papers or [], 1):
        idx = p.get("idx")
        e = extractions_by_idx.get(idx) or {}
        included = included_map.get(idx, False)
        concepts = e.get("concepts") or []
        if isinstance(concepts, str):
            concepts = [concepts]
        rows.append([
            i, p.get("title") or "", p.get("authors") or "", p.get("year") or "",
            p.get("venue") or "", p.get("source") or "", p.get("url") or "",
            "Yes" if included else "No", p.get("abstract") or "",
            e.get("method") or "", e.get("finding") or "", e.get("data") or "",
            e.get("metrics") or "", e.get("limitation") or "", e.get("contribution") or "",
            e.get("relevance") or "", ", ".join(concepts),
        ])
    return rows


def papers_to_csv(papers: list[dict], extractions_by_idx: dict, included_map: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    w.writerows(_rows(papers, extractions_by_idx, included_map))
    # utf-8-sig so Excel (which otherwise guesses the wrong encoding on a
    # plain utf-8 CSV) opens accented author names / non-ASCII titles correctly.
    return buf.getvalue().encode("utf-8-sig")


def papers_to_xlsx(papers: list[dict], extractions_by_idx: dict, included_map: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="6D5EF6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in _rows(papers, extractions_by_idx, included_map):
        ws.append(row)

    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
